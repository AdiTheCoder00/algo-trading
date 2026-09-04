"""The Excel bridge, exercised with no Excel and no broker.

Everything interesting about this feature is a decision made before a cell is
written: which contract a typed symbol means, whether an armed row is safe to send,
what a missing book looks like, and — the one that matters most — whether an order
row can ever be sent twice. `SheetIO` and the broker are both protocols, so all of
that is reachable from a plain test.

The order tests are deliberately the fussiest ones here. A quote written into the
wrong cell is visible; an order sent twice is not.
"""

from __future__ import annotations

from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import pytest

from algo.core.enums import Exchange, Mode, OrderType, ProductType, Right, Side
from algo.core.errors import DataError
from algo.core.order import BrokerOrderRef, Order
from algo.core.quote import Quote
from algo.excel import layout, orders, sync
from algo.excel.io import create_workbook, verify_layout
from algo.excel.service import ExcelBridge
from algo.exchange.master import InstrumentMaster, MasterRow
from algo.execution.broker import BrokerHealth, BrokerPositionSnapshot, Funds

NOW = datetime(2026, 9, 1, 6, 30, tzinfo=UTC)
FRONT_EXPIRY = date(2026, 9, 30)
BACK_EXPIRY = date(2026, 10, 31)
OPTION_EXPIRY = date(2026, 9, 25)


class FixedClock:
    def now(self) -> datetime:
        return NOW


class FakeSheetIO:
    """A grid per sheet that grows on demand, recording the order of writes.

    `writes` is what the order tests assert against: the safety property is not
    "STATUS ends up correct" but "STATUS was written before the broker was called",
    and only a log of the sequence can show that.
    """

    def __init__(self, sheets: dict[str, list[list[Any]]] | None = None) -> None:
        self.grids: dict[str, list[list[Any]]] = sheets or {}
        self.writes: list[tuple[str, int, int, list[list[Any]]]] = []

    def _grid(self, sheet: str) -> list[list[Any]]:
        return self.grids.setdefault(sheet, [])

    def _ensure(self, sheet: str, row: int, col: int) -> None:
        grid = self._grid(sheet)
        while len(grid) < row:
            grid.append([])
        for line in grid:
            while len(line) < col:
                line.append(None)

    def read(self, sheet: str, row: int, col: int, n_rows: int, n_cols: int) -> list[list[Any]]:
        self._ensure(sheet, row + n_rows - 1, col + n_cols - 1)
        grid = self._grid(sheet)
        return [
            [grid[r][c] for c in range(col - 1, col - 1 + n_cols)]
            for r in range(row - 1, row - 1 + n_rows)
        ]

    def write(self, sheet: str, row: int, col: int, values: Any) -> None:
        if not values:
            return
        rows = [list(line) for line in values]
        self.writes.append((sheet, row, col, rows))
        self._ensure(sheet, row + len(rows) - 1, col + max(len(r) for r in rows) - 1)
        grid = self._grid(sheet)
        for r_offset, line in enumerate(rows):
            for c_offset, value in enumerate(line):
                grid[row - 1 + r_offset][col - 1 + c_offset] = value

    def cell(self, sheet: str, row: int, col: int) -> Any:
        return self.read(sheet, row, col, 1, 1)[0][0]


def _master() -> InstrumentMaster:
    """Two futures months and a call/put pair, which is enough to make every
    resolution rule observable: front versus back month, and exact versus derived."""
    rows = [
        MasterRow(
            symboltoken="1",
            tradingsymbol="GOLDM25SEPFUT",
            exch_seg="MCX",
            name="GOLDM",
            instrumenttype="FUTCOM",
            expiry=FRONT_EXPIRY,
            lot_size=Decimal("10"),
        ),
        MasterRow(
            symboltoken="2",
            tradingsymbol="GOLDM25OCTFUT",
            exch_seg="MCX",
            name="GOLDM",
            instrumenttype="FUTCOM",
            expiry=BACK_EXPIRY,
            lot_size=Decimal("10"),
        ),
        MasterRow(
            symboltoken="3",
            tradingsymbol="GOLDM25SEP150000CE",
            exch_seg="MCX",
            name="GOLDM",
            instrumenttype="OPTFUT",
            expiry=OPTION_EXPIRY,
            strike=Decimal("150000"),
            lot_size=Decimal("10"),
        ),
        MasterRow(
            symboltoken="4",
            tradingsymbol="GOLDM25SEP150000PE",
            exch_seg="MCX",
            name="GOLDM",
            instrumenttype="OPTFUT",
            expiry=OPTION_EXPIRY,
            strike=Decimal("150000"),
            lot_size=Decimal("10"),
        ),
    ]
    return InstrumentMaster(rows, fetched_at=NOW)


def _quote(**overrides: Any) -> Quote:
    fields: dict[str, Any] = {
        "exchange_ts": NOW,
        "received_ts": NOW,
        "bid": Decimal("149990"),
        "ask": Decimal("150010"),
        "bid_qty": 5,
        "ask_qty": 7,
        "ltp": Decimal("150000"),
        "volume": 120,
        "open_interest": 900,
    }
    fields.update(overrides)
    return Quote(**fields)


# --------------------------------------------------------------- resolution
def test_exact_tradingsymbol_beats_the_underlying() -> None:
    """The back month must be reachable by name even though GOLDM alone is front."""
    row = sync.resolve_row(_master(), "GOLDM25OCTFUT")
    assert row.tradingsymbol == "GOLDM25OCTFUT"
    assert row.expiry == BACK_EXPIRY


def test_bare_underlying_resolves_to_the_front_month() -> None:
    assert sync.resolve_row(_master(), "goldm").expiry == FRONT_EXPIRY


def test_unknown_symbol_names_itself_in_the_error() -> None:
    with pytest.raises(DataError, match="NOTHING"):
        sync.resolve_row(_master(), "nothing")


# ------------------------------------------------------------------- quotes
def test_a_missing_bid_stays_empty_rather_than_zero() -> None:
    """The whole tradeability distinction rests on this: "nobody is quoting" and
    "the quote is zero" must not render the same."""
    cells = sync.quote_cells(_master().future_rows("GOLDM", Exchange.MCX)[0],
                             _quote(bid=None, bid_qty=None), now_ist=NOW)
    assert cells[2] is None  # BID
    assert cells[4] is None  # BID_QTY


def test_a_row_with_no_quote_at_all_says_so() -> None:
    cells = sync.quote_cells(
        _master().future_rows("GOLDM", Exchange.MCX)[0], None, now_ist=NOW
    )
    assert cells[-2] == "NO QUOTE"
    assert cells[1] is None  # LTP


# -------------------------------------------------------------------- chain
def test_the_ladder_mirrors_puts_and_keeps_one_sided_strikes() -> None:
    """A strike quoted only as a call still gets a row. Dropping it would move the
    apparent edge of the chain."""
    from algo.core.chain import ChainRow, OptionChainSnapshot
    from algo.core.instrument import FutureId, OptionId

    def option(strike: str, right: Right) -> OptionId:
        return OptionId(
            underlying_future=FutureId(underlying="GOLDM", expiry=FRONT_EXPIRY),
            option_expiry=OPTION_EXPIRY,
            strike=Decimal(strike),
            right=right,
        )

    snapshot = OptionChainSnapshot(
        ts=NOW,
        underlying="GOLDM",
        option_expiry=OPTION_EXPIRY,
        futures_price=Decimal("150000"),
        futures_quote=_quote(),
        rows=(
            ChainRow(option=option("150000", Right.CE), quote=_quote(ltp=Decimal("11"))),
            ChainRow(option=option("150000", Right.PE), quote=_quote(ltp=Decimal("22"))),
            ChainRow(option=option("151000", Right.CE), quote=_quote(ltp=Decimal("33"))),
        ),
    )
    table = sync.chain_cells(snapshot)
    assert len(table) == 2

    first = table[0]
    assert first[5] == 150000.0  # STRIKE sits in the middle
    assert first[4] == 11.0  # CE_LTP is the column beside it
    assert first[6] == 22.0  # PE_LTP mirrors it on the other side

    one_sided = table[1]
    assert one_sided[4] == 33.0  # the call is there
    assert one_sided[6] is None  # the absent put is blank, not dropped


# ---------------------------------------------------------------- portfolio
def test_a_short_position_in_profit_reports_a_positive_pnl() -> None:
    """Signed qty carries the direction, so the multiplication must not be
    branched on side — this is the test that a short reading as a loss would fail."""
    rows = sync.position_cells(
        [
            BrokerPositionSnapshot(
                instrument_key="MCX:GOLDM:FUT:20260930",
                qty=Decimal("-20"),
                lots=-2,
                average_price=Decimal("150000"),
            )
        ],
        {"MCX:GOLDM:FUT:20260930": Decimal("149000")},
    )
    assert rows[0][5] == pytest.approx(20000.0)


def test_a_position_with_no_mark_reports_no_pnl() -> None:
    rows = sync.position_cells(
        [
            BrokerPositionSnapshot(
                instrument_key="MCX:GOLDM:FUT:20260930",
                qty=Decimal("10"),
                lots=1,
                average_price=Decimal("150000"),
            )
        ],
        {},
    )
    assert rows[0][4] is None and rows[0][5] is None


def test_pad_truncates_wide_rows_and_fills_short_tables() -> None:
    padded = sync.pad([[1, 2, 3, 4]], to_rows=3, width=2)
    assert padded == [[1, 2], [None, None], [None, None]]


# ------------------------------------------------------------------- orders
def _order_row(**overrides: Any) -> list[Any]:
    row: dict[str, Any] = {
        "symbol": "GOLDM",
        "side": "SELL",
        "lots": 2,
        "order_type": "LIMIT",
        "product": "NRML",
        "limit": 150000,
        "trigger": None,
        "send": "YES",
        "status": None,
    }
    row.update(overrides)
    return [
        row["symbol"],
        row["side"],
        row["lots"],
        row["order_type"],
        row["product"],
        row["limit"],
        row["trigger"],
        row["send"],
        row["status"],
        None,
        None,
        None,
    ]


def test_a_row_is_armed_only_by_an_explicit_send_value() -> None:
    assert orders.is_armed(_order_row(send="yes"))
    assert not orders.is_armed(_order_row(send=""))
    assert not orders.is_armed(_order_row(send="maybe"))


def test_a_row_already_stamped_is_never_armed_again() -> None:
    """The interlock. Without it every completed row re-fires on the next tick."""
    assert not orders.is_armed(_order_row(send="YES", status="SENT"))
    assert not orders.is_armed(_order_row(send="YES", status=orders.SENDING))


def test_a_market_order_carrying_a_limit_price_is_refused() -> None:
    with pytest.raises(DataError, match="must not carry a LIMIT_PRICE"):
        orders.parse_row(_order_row(order_type="MARKET", limit=150000), row=2)


def test_a_limit_order_without_a_price_is_refused() -> None:
    with pytest.raises(DataError, match="needs a LIMIT_PRICE"):
        orders.parse_row(_order_row(order_type="LIMIT", limit=None), row=2)


def test_an_unreadable_side_names_the_column() -> None:
    with pytest.raises(DataError, match="SIDE"):
        orders.parse_row(_order_row(side="BUYY"), row=2)


def test_zero_lots_is_refused() -> None:
    with pytest.raises(DataError, match="at least one lot"):
        orders.parse_row(_order_row(lots=0), row=2)


def test_a_blank_product_defaults_to_nrml() -> None:
    armed = orders.parse_row(_order_row(product=None), row=2)
    assert armed is not None and armed.product is ProductType.NRML


def test_lots_become_quantity_via_the_master_lot_size() -> None:
    """The operator trades in lots; a hand-typed quantity is a decimal point away
    from being a hundred times too large."""
    master = _master()
    armed = orders.parse_row(_order_row(lots=3), row=2)
    assert armed is not None
    order = orders.build_order(
        armed,
        sync.resolve_row(master, "GOLDM"),
        master,
        now=NOW,
        order_id="XLtest",
    )
    assert order.lots == 3
    assert order.qty == Decimal("30")


def test_an_option_order_does_not_conflate_the_two_expiries() -> None:
    """`instrument.py` is emphatic that the option and its underlying future expire
    on different dates, and conflating them is what walks a short leg into
    devolvement."""
    from algo.core.instrument import OptionId

    master = _master()
    row = sync.resolve_row(master, "GOLDM25SEP150000CE")
    instrument = orders.instrument_for(row, master)
    assert isinstance(instrument, OptionId)
    assert instrument.option_expiry == OPTION_EXPIRY
    assert instrument.underlying_future.expiry == FRONT_EXPIRY
    assert instrument.option_expiry != instrument.underlying_future.expiry


@pytest.mark.parametrize("mode", [Mode.PAPER, Mode.BACKTEST])
def test_every_mode_but_live_refuses_to_send(mode: Mode) -> None:
    reason = orders.gate_reason(mode)
    assert reason is not None
    assert "not sent" in reason
    # The reason has to name all three conditions, because the failure to avoid is
    # an operator disabling the check out of frustration at an unclear message.
    assert "--mode live" in reason
    assert "TRADING_MODE=live" in reason
    assert "--i-understand-this-is-real-money" in reason


def test_live_mode_is_the_only_one_that_sends() -> None:
    assert orders.gate_reason(Mode.LIVE) is None


# ------------------------------------------------------------------ service
class FakeQuotes:
    def __init__(self, payloads: list[dict[str, Any]] | None = None) -> None:
        self.payloads = payloads if payloads is not None else []
        self.calls: list[list[dict[str, str]]] = []

    def quotes(self, exchange_segments: list[dict[str, str]]) -> Any:
        self.calls.append(exchange_segments)
        wanted = {entry["instrument_token"] for entry in exchange_segments}
        return [p for p in self.payloads if str(p.get("exchange_token")) in wanted]


class RecordingBroker:
    """Records the sheet as it looked at the moment `place` was called."""

    def __init__(self, io: FakeSheetIO, *, fail: bool = False) -> None:
        self._io = io
        self._fail = fail
        self.placed: list[Order] = []
        self.status_when_called: list[Any] = []

    def place(self, order: Order) -> BrokerOrderRef:
        self.placed.append(order)
        self.status_when_called.append(
            self._io.cell(layout.Orders.SHEET, layout.FIRST_DATA_ROW, layout.Orders.STATUS_COL)
        )
        if self._fail:
            raise DataError("broker said no")
        return BrokerOrderRef(
            client_order_id=order.client_order_id,
            broker_order_id="NEO-1",
            accepted_at=NOW,
        )

    def positions(self) -> list[BrokerPositionSnapshot]:
        return []

    def funds(self) -> Funds:
        return Funds(cash=Decimal("100000"))

    def health(self) -> BrokerHealth:
        return BrokerHealth(connected=True, last_heartbeat=NOW, detail="fake session")


def _io_with_one_order(**overrides: Any) -> FakeSheetIO:
    io = FakeSheetIO()
    io.write(layout.Orders.SHEET, layout.FIRST_DATA_ROW, 1, [_order_row(**overrides)])
    io.writes.clear()
    return io


def _bridge(io: FakeSheetIO, *, mode: Mode, broker: Any = None) -> ExcelBridge:
    return ExcelBridge(
        io=io,
        master=_master(),
        quotes=FakeQuotes(),
        clock=FixedClock(),
        mode=mode,
        broker=broker,
    )


def test_a_dry_run_validates_the_row_and_never_touches_the_broker() -> None:
    io = _io_with_one_order()
    broker = RecordingBroker(io)
    _bridge(io, mode=Mode.PAPER, broker=broker).refresh()

    assert broker.placed == []
    status = io.cell(layout.Orders.SHEET, layout.FIRST_DATA_ROW, layout.Orders.STATUS_COL)
    assert str(status).startswith("DRY RUN")


def test_a_dry_run_still_reports_a_bad_symbol() -> None:
    """Validating without sending is the point of paper mode; a run that accepted
    anything would teach the operator nothing before they went live."""
    io = _io_with_one_order(symbol="NOSUCH")
    _bridge(io, mode=Mode.PAPER, broker=RecordingBroker(io)).refresh()
    status = str(io.cell(layout.Orders.SHEET, layout.FIRST_DATA_ROW, layout.Orders.STATUS_COL))
    assert status.startswith("REJECTED")


def test_status_is_stamped_before_the_broker_is_called() -> None:
    """The property that makes a crash mid-send safe. If STATUS were written after
    the call, a crash in the window would leave a blank row that the next tick
    would send again."""
    io = _io_with_one_order()
    broker = RecordingBroker(io)
    _bridge(io, mode=Mode.LIVE, broker=broker).refresh()

    assert broker.status_when_called == [orders.SENDING]
    assert io.cell(layout.Orders.SHEET, layout.FIRST_DATA_ROW, layout.Orders.STATUS_COL) == "SENT"
    assert (
        io.cell(layout.Orders.SHEET, layout.FIRST_DATA_ROW, layout.Orders.STATUS_COL + 2)
        == "NEO-1"
    )


def test_the_same_row_is_not_sent_twice_across_ticks() -> None:
    io = _io_with_one_order()
    broker = RecordingBroker(io)
    bridge = _bridge(io, mode=Mode.LIVE, broker=broker)
    bridge.refresh()
    bridge.refresh()
    bridge.refresh()
    assert len(broker.placed) == 1


def test_a_rejected_order_says_so_and_is_not_retried() -> None:
    io = _io_with_one_order()
    broker = RecordingBroker(io, fail=True)
    bridge = _bridge(io, mode=Mode.LIVE, broker=broker)
    bridge.refresh()
    bridge.refresh()

    assert len(broker.placed) == 1  # a failure is an outcome, not a reason to retry
    status = str(io.cell(layout.Orders.SHEET, layout.FIRST_DATA_ROW, layout.Orders.STATUS_COL))
    assert status.startswith("FAILED")


def test_live_mode_without_a_session_refuses_rather_than_pretending() -> None:
    io = _io_with_one_order()
    _bridge(io, mode=Mode.LIVE, broker=None).refresh()
    status = str(io.cell(layout.Orders.SHEET, layout.FIRST_DATA_ROW, layout.Orders.STATUS_COL))
    assert status.startswith("NO SESSION")


def test_the_order_id_is_stable_within_a_second_and_prefixed() -> None:
    armed = orders.parse_row(_order_row(), row=2)
    assert armed is not None
    first = orders.client_order_id(armed, at=NOW)
    assert first == orders.client_order_id(armed, at=NOW.replace(microsecond=500))
    assert first.startswith("XL")
    assert orders.client_order_id(armed, at=NOW) != orders.client_order_id(
        orders.parse_row(_order_row(lots=5), row=2) or armed, at=NOW
    )


def test_quotes_reach_the_sheet_and_become_the_portfolio_mark() -> None:
    io = FakeSheetIO()
    io.write(layout.Quotes.SHEET, layout.FIRST_DATA_ROW, 1, [["GOLDM"]])
    quotes = FakeQuotes(
        [
            {
                "exchange_token": "1",
                "ltp": "150000",
                "volume": 120,
                "open_int": 900,
                "depth": {
                    "buy": [{"price": "149990", "quantity": 5}],
                    "sell": [{"price": "150010", "quantity": 7}],
                },
            }
        ]
    )
    bridge = ExcelBridge(
        io=io,
        master=_master(),
        quotes=quotes,
        clock=FixedClock(),
        mode=Mode.PAPER,
    )
    bridge.refresh()

    assert io.cell(layout.Quotes.SHEET, layout.FIRST_DATA_ROW, 2) == "GOLDM25SEPFUT"
    assert io.cell(layout.Quotes.SHEET, layout.FIRST_DATA_ROW, 3) == 150000.0
    assert io.cell(layout.Quotes.SHEET, layout.FIRST_DATA_ROW, 4) == 149990.0
    assert io.cell(layout.Status.SHEET, 5, layout.Status.VALUE_COL) == ""  # no LAST_ERROR


def test_a_failing_section_is_reported_without_stopping_the_others() -> None:
    """A market-data hiccup must not stop the order sheet from being drained."""

    class BrokenQuotes:
        def quotes(self, exchange_segments: list[dict[str, str]]) -> Any:
            raise DataError("feed down")

    io = _io_with_one_order()
    io.write(layout.Quotes.SHEET, layout.FIRST_DATA_ROW, 1, [["GOLDM"]])
    broker = RecordingBroker(io)
    ExcelBridge(
        io=io,
        master=_master(),
        quotes=BrokenQuotes(),
        clock=FixedClock(),
        mode=Mode.LIVE,
        broker=broker,
    ).refresh()

    assert len(broker.placed) == 1  # orders still drained
    assert "feed down" in str(io.cell(layout.Status.SHEET, 5, layout.Status.VALUE_COL))


# ------------------------------------------------------------------- layout
def test_a_created_workbook_passes_its_own_layout_check(tmp_path: Path) -> None:
    """The scaffolder and the verifier must agree, or every fresh workbook is
    rejected on first attach."""
    import openpyxl

    path = create_workbook(tmp_path / "bridge.xlsx")
    book = openpyxl.load_workbook(path)

    class OpenpyxlIO:
        def read(
            self, sheet: str, row: int, col: int, n_rows: int, n_cols: int
        ) -> list[list[Any]]:
            ws = book[sheet]
            return [
                [ws.cell(row=r, column=c).value for c in range(col, col + n_cols)]
                for r in range(row, row + n_rows)
            ]

        def write(self, sheet: str, row: int, col: int, values: Any) -> None:
            raise AssertionError("verify_layout must not write")

    verify_layout(OpenpyxlIO())


def test_a_moved_column_is_refused_with_both_layouts_shown(tmp_path: Path) -> None:
    io = FakeSheetIO()
    for sheet, header_row, headers in (
        (layout.Quotes.SHEET, layout.HEADER_ROW, layout.Quotes.HEADERS),
        (layout.Chain.SHEET, layout.Chain.HEADER_ROW, layout.Chain.HEADERS),
        (layout.Orders.SHEET, layout.HEADER_ROW, layout.Orders.HEADERS),
        (layout.Portfolio.SHEET, layout.Portfolio.HEADER_ROW, layout.Portfolio.HEADERS),
    ):
        io.write(sheet, header_row, 1, [list(headers)])
    # Swap SIDE and LOTS — the exact drift that would otherwise place an order for
    # a quantity read out of the side column.
    io.write(layout.Orders.SHEET, layout.HEADER_ROW, 2, [["LOTS", "SIDE"]])

    with pytest.raises(DataError, match="does not match the expected layout"):
        verify_layout(io)


def test_creating_over_an_existing_workbook_needs_permission(tmp_path: Path) -> None:
    path = create_workbook(tmp_path / "bridge.xlsx")
    with pytest.raises(DataError, match="already exists"):
        create_workbook(path)
    assert create_workbook(path, overwrite=True) == path


def test_a_blank_expiry_means_the_nearest_listed_one() -> None:
    assert sync.parse_expiry(None) is None
    assert sync.parse_expiry("") is None
    assert sync.parse_expiry("2026-09-25") == OPTION_EXPIRY
    assert sync.parse_expiry(datetime(2026, 9, 25, 12, 0)) == OPTION_EXPIRY


def test_an_unreadable_expiry_says_what_it_wanted() -> None:
    with pytest.raises(DataError, match="YYYY-MM-DD"):
        sync.parse_expiry("next month")


def test_side_is_read_as_an_enum_not_a_string() -> None:
    armed = orders.parse_row(_order_row(side="sell"), row=2)
    assert armed is not None
    assert armed.side is Side.SELL
    assert armed.order_type is OrderType.LIMIT


# --------------------------------------------------------------- chain sheet
def _chain_master() -> InstrumentMaster:
    """The shared master plus quotes for its tokens is enough for a real feed."""
    return _master()


def _chain_payloads() -> list[dict[str, Any]]:
    """Deliberately asymmetric bids and asks.

    A fixture where both sides quote the same numbers cannot tell a correct
    ladder from one that has swapped the put book, which is exactly how the
    put-side bid/ask transposition survived the first round of tests.
    """
    def q(token: str, *, ltp: str, bid: str, ask: str) -> dict[str, Any]:
        return {
            "exchange_token": token,
            "ltp": ltp,
            "open_int": "2000",
            "volume": 55,
            "depth": {
                "buy": [{"price": bid, "quantity": "3"}],
                "sell": [{"price": ask, "quantity": "4"}],
            },
        }

    return [
        q("1", ltp="150000", bid="149990", ask="150010"),  # the anchor future
        q("3", ltp="700", bid="690", ask="710"),  # 150000 CE
        q("4", ltp="500", bid="480", ask="520"),  # 150000 PE
    ]


def _feed_factory(master: InstrumentMaster, quotes: FakeQuotes, seen: list[str]) -> Any:
    """A real `KotakChainFeed` per underlying, recording what it was asked for."""
    from algo.data.kotak_feed import KotakChainFeed
    from algo.data.live import SessionWindow
    from algo.exchange.calendar import synthetic_calendar

    def make(underlying: str) -> KotakChainFeed:
        seen.append(underlying)
        return KotakChainFeed(
            transport=quotes,
            master=master,
            underlying=underlying,
            clock=FixedClock(),
            session=SessionWindow(synthetic_calendar()),
            poll_interval_s=0.0,
        )

    return make


def _chain_bridge(io: FakeSheetIO, seen: list[str]) -> ExcelBridge:
    master = _chain_master()
    quotes = FakeQuotes(_chain_payloads())
    return ExcelBridge(
        io=io,
        master=master,
        quotes=quotes,
        clock=FixedClock(),
        mode=Mode.PAPER,
        chain_feed_for=_feed_factory(master, quotes, seen),
    )


def test_the_put_book_is_not_transposed_by_the_mirrored_layout() -> None:
    """The put columns mirror the calls in ORDER only; a bid is still a bid.

    Reversing the call list to build them put ASK under PE_BID and BID under
    PE_ASK, so every put read as bid above its own offer — a crossed book, which
    is precisely the shape a tradeability check treats as free money.
    """
    io = FakeSheetIO()
    io.write(layout.Chain.SHEET, *layout.Chain.UNDERLYING_CELL, [["GOLDM"]])
    _chain_bridge(io, []).refresh()

    row = io.read(layout.Chain.SHEET, layout.Chain.FIRST_DATA_ROW, 1, 1, 11)[0]
    names = dict(zip(layout.Chain.HEADERS, row, strict=True))

    assert names["STRIKE"] == 150000.0
    assert names["CE_BID"] == 690.0 and names["CE_ASK"] == 710.0
    assert names["PE_BID"] == 480.0 and names["PE_ASK"] == 520.0
    # The invariant underneath the assertions above: neither side is crossed.
    assert names["CE_BID"] < names["CE_ASK"]
    assert names["PE_BID"] < names["PE_ASK"]
    assert names["PE_LTP"] == 500.0 and names["CE_LTP"] == 700.0


def test_the_chain_writes_the_futures_anchor_above_the_ladder() -> None:
    io = FakeSheetIO()
    io.write(layout.Chain.SHEET, *layout.Chain.UNDERLYING_CELL, [["goldm"]])
    _chain_bridge(io, []).refresh()

    assert io.cell(layout.Chain.SHEET, *layout.Chain.FUTURES_PRICE_CELL) == 150000.0
    assert io.cell(layout.Status.SHEET, 5, layout.Status.VALUE_COL) == ""  # no LAST_ERROR


def test_a_blank_expiry_falls_back_to_the_nearest_listed_one() -> None:
    """Leaving the cell empty is not an error - it means "the front cycle"."""
    io = FakeSheetIO()
    io.write(layout.Chain.SHEET, *layout.Chain.UNDERLYING_CELL, [["GOLDM"]])
    io.write(layout.Chain.SHEET, *layout.Chain.EXPIRY_CELL, [[None]])
    _chain_bridge(io, []).refresh()

    # OPTION_EXPIRY is the only listed cycle, so a ladder proves it was chosen.
    assert io.cell(layout.Chain.SHEET, layout.Chain.FIRST_DATA_ROW, 6) == 150000.0


def test_an_explicit_expiry_is_honoured() -> None:
    io = FakeSheetIO()
    io.write(layout.Chain.SHEET, *layout.Chain.UNDERLYING_CELL, [["GOLDM"]])
    io.write(layout.Chain.SHEET, *layout.Chain.EXPIRY_CELL, [[OPTION_EXPIRY.isoformat()]])
    _chain_bridge(io, []).refresh()

    assert io.cell(layout.Chain.SHEET, layout.Chain.FIRST_DATA_ROW, 6) == 150000.0


def test_a_blank_underlying_leaves_the_chain_untouched() -> None:
    """No underlying means the operator has not asked for a chain, which is not
    the same as asking for an empty one - the feed must not even be built."""
    io = FakeSheetIO()
    seen: list[str] = []
    _chain_bridge(io, seen).refresh()

    assert seen == []
    assert io.cell(layout.Chain.SHEET, *layout.Chain.FUTURES_PRICE_CELL) is None


def test_an_underlying_with_no_options_is_reported_not_raised() -> None:
    """A bad chain must not stop the tick: quotes and orders still run, and the
    reason lands in Status where the operator will actually see it."""
    io = FakeSheetIO()
    io.write(layout.Chain.SHEET, *layout.Chain.UNDERLYING_CELL, [["NOSUCH"]])
    _chain_bridge(io, []).refresh()

    error = str(io.cell(layout.Status.SHEET, 5, layout.Status.VALUE_COL))
    assert "chain:" in error
    assert "NOSUCH" in error
