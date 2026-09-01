"""The boundary between the bridge and Excel itself.

Everything above this module manipulates lists of cell values. Only `XlwingsSheetIO`
knows that a workbook is a running copy of Excel — which is what lets the whole
bridge, order parsing included, be tested on a machine with no Excel installed, the
same way `KotakTransport` lets the broker be tested with no broker.

**Why a live workbook and not a written file.** `openpyxl` can only produce a file
on disk; it cannot see what the operator has typed into a workbook that is open in
front of them, and saving over an open file does not update it. An order sheet has
to be read while the operator is using it, so the live path goes through COM.
`create_workbook` still uses openpyxl, because scaffolding a new file is the one
job that genuinely has no reader.

**Batching.** Every read and write is one rectangular range, because each COM call
is an out-of-process round trip. Filling 200 rows cell by cell takes seconds; doing
it as one range assignment takes milliseconds.
"""

from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path
from typing import Any, Protocol, runtime_checkable

from algo.core.errors import DataError
from algo.excel import layout


@runtime_checkable
class SheetIO(Protocol):
    """The workbook surface the bridge needs. Fakes implement the same protocol."""

    def read(self, sheet: str, row: int, col: int, n_rows: int, n_cols: int) -> list[list[Any]]:
        """Read a rectangle, 1-based and inclusive. Empty cells come back as None."""
        ...

    def write(self, sheet: str, row: int, col: int, values: Sequence[Sequence[Any]]) -> None:
        """Write a rectangle whose top-left is (row, col). A no-op for an empty
        `values`, so callers need not special-case an empty table."""
        ...


class XlwingsSheetIO:
    """`SheetIO` against a workbook open in Excel, via xlwings/COM.

    xlwings is an optional, Windows-only extra (`pip install -e ".[excel]"`) for the
    same reason MetaTrader5 is: there is no wheel for Linux, and putting it in the
    base dependencies would break the CI install. `algo/` never imports it except
    here, behind `SheetIO`, so the type and test gates run without it.
    """

    __slots__ = ("_book",)

    def __init__(self, book: Any) -> None:
        self._book = book

    @classmethod
    def attach(cls, path: Path | str, *, visible: bool = True) -> XlwingsSheetIO:
        """Attach to `path` if Excel already has it open, otherwise open it.

        Attaching rather than always opening is what lets the operator keep the
        workbook in front of them while the bridge runs. Opening a second copy of an
        already-open workbook would give the bridge a private view whose order rows
        nobody is typing into.
        """
        import xlwings  # imported here so the package imports without Excel

        target = Path(path).resolve()
        for book in xlwings.books:
            try:
                if Path(book.fullname).resolve() == target:
                    return cls(book)
            except (OSError, ValueError):  # a never-saved book has no usable path
                continue
        # `apps.active` is None when Excel is running but no window owns the
        # foreground, which is normal under a scheduler; start our own then.
        app = xlwings.apps.active if xlwings.apps else None
        if app is None:
            app = xlwings.App(visible=visible)
        return cls(app.books.open(str(target)))

    def read(self, sheet: str, row: int, col: int, n_rows: int, n_cols: int) -> list[list[Any]]:
        rng = self._book.sheets[sheet].range((row, col), (row + n_rows - 1, col + n_cols - 1))
        value = rng.value
        # xlwings collapses a 1x1 range to a scalar and a 1xN or Nx1 range to a flat
        # list. Normalising here keeps every caller working in rectangles.
        if n_rows == 1 and n_cols == 1:
            return [[value]]
        if n_rows == 1:
            return [list(value)]
        if n_cols == 1:
            return [[cell] for cell in value]
        return [list(line) for line in value]

    def write(self, sheet: str, row: int, col: int, values: Sequence[Sequence[Any]]) -> None:
        if not values:
            return
        self._book.sheets[sheet].range((row, col)).value = [list(line) for line in values]

    def save(self) -> None:
        self._book.save()


def verify_layout(io: SheetIO) -> None:
    """Check every header the bridge writes against, and say precisely what moved.

    Called once on attach rather than per refresh. The failure this prevents is the
    quiet one: a column inserted into `Orders` would otherwise make the bridge read
    LOTS out of the SIDE column and place an order nobody asked for.
    """
    checks = (
        (layout.Quotes.SHEET, layout.HEADER_ROW, layout.Quotes.HEADERS),
        (layout.Chain.SHEET, layout.Chain.HEADER_ROW, layout.Chain.HEADERS),
        (layout.Orders.SHEET, layout.HEADER_ROW, layout.Orders.HEADERS),
        (layout.Portfolio.SHEET, layout.Portfolio.HEADER_ROW, layout.Portfolio.HEADERS),
    )
    for sheet, header_row, expected in checks:
        try:
            found_raw = io.read(sheet, header_row, 1, 1, len(expected))[0]
        except Exception as exc:
            raise DataError(
                f"cannot read the header of sheet {sheet!r}: {exc}. "
                "Rebuild the workbook with --recreate."
            ) from exc
        found = tuple("" if cell is None else str(cell).strip() for cell in found_raw)
        if found != expected:
            raise DataError(
                f"sheet {sheet!r} row {header_row} does not match the expected layout.\n"
                f"  expected: {list(expected)}\n"
                f"  found:    {list(found)}\n"
                "Columns are read by position, so the bridge refuses to write into a "
                "sheet it no longer recognises. Rebuild it with --recreate."
            )


def create_workbook(path: Path | str, *, overwrite: bool = False) -> Path:
    """Scaffold a fresh workbook with every sheet and header in place.

    Uses openpyxl, so a workbook can be created on a machine with no Excel — the
    file is the deliverable here, and there is no open copy to disturb.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    target = Path(path)
    if target.exists() and not overwrite:
        raise DataError(
            f"{target} already exists. Pass --recreate to overwrite it — doing so "
            "discards whatever is currently in its Orders sheet."
        )
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    default = workbook.active
    assert default is not None
    workbook.remove(default)

    bold = Font(name="Calibri", bold=True)
    header_font = Font(name="Calibri", bold=True)

    def header(sheet: Any, row: int, headers: tuple[str, ...]) -> None:
        for index, text in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=index, value=text)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            sheet.column_dimensions[get_column_letter(index)].width = max(12, len(text) + 2)

    status = workbook.create_sheet(layout.Status.SHEET)
    for offset, label in enumerate(layout.Status.LABELS):
        status.cell(
            row=layout.Status.FIRST_ROW + offset,
            column=layout.Status.LABEL_COL,
            value=label,
        ).font = bold
    status.column_dimensions["A"].width = 20
    status.column_dimensions["B"].width = 60

    quotes = workbook.create_sheet(layout.Quotes.SHEET)
    header(quotes, layout.HEADER_ROW, layout.Quotes.HEADERS)
    # A worked example, so the first run shows something rather than nothing.
    quotes.cell(row=layout.FIRST_DATA_ROW, column=1, value="GOLDM")

    chain = workbook.create_sheet(layout.Chain.SHEET)
    chain.cell(row=1, column=1, value="UNDERLYING").font = bold
    chain.cell(row=2, column=1, value="EXPIRY").font = bold
    chain.cell(row=3, column=1, value="FUTURES_PRICE").font = bold
    chain.cell(
        row=layout.Chain.UNDERLYING_CELL[0],
        column=layout.Chain.UNDERLYING_CELL[1],
        value="GOLDM",
    )
    header(chain, layout.Chain.HEADER_ROW, layout.Chain.HEADERS)
    chain.column_dimensions["A"].width = 16

    orders = workbook.create_sheet(layout.Orders.SHEET)
    header(orders, layout.HEADER_ROW, layout.Orders.HEADERS)

    portfolio = workbook.create_sheet(layout.Portfolio.SHEET)
    portfolio.cell(row=1, column=1, value="CASH").font = bold
    portfolio.cell(row=2, column=1, value="MARGIN_USED").font = bold
    portfolio.cell(row=3, column=1, value="MARGIN_AVAILABLE").font = bold
    header(portfolio, layout.Portfolio.HEADER_ROW, layout.Portfolio.HEADERS)
    portfolio.column_dimensions["A"].width = 30

    workbook.save(target)
    workbook.close()
    return target
