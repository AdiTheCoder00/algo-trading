"""The live MCX option chain, as scraped from the exchange site into Excel.

This is the only source in the project that carries a **real book**. Bhavcopy is
end-of-day and has no bid or ask at all, so every tradeability decision made
against it rests on `assume_spread`'s invention; SmartAPI can only serve
contracts that have not yet expired. A scrape of the live chain page has what
both lack: a genuine bid, ask, and depth-at-top for every listed strike, at one
instant.

What it does not have is history. This is one snapshot, not a series, so it can
show what the ladder looks like and what is actually quotable right now — it
cannot backtest anything. The two sources answer different questions and neither
replaces the other.

## The layout is a rendered page, not a data format

The sheet mirrors the exchange's on-screen ladder: calls on the left, strike in
the middle, puts on the right, with three title lines above the header. That
means the column *positions* carry the meaning and the header text repeats
itself ("LTP" appears twice, once per side), so this reads by position against a
verified header row rather than by name. `_verify_header` fails loudly if the
scrape's layout shifts, for the same reason `bhavcopy.parse_rows` does: a
silently mis-mapped column produces a chain that looks right and prices wrong.

Blank bid/ask cells are preserved as None, never coerced to zero. "Nobody is
quoting this strike" and "the quote is zero" are different facts, and the whole
tradeability gate depends on telling them apart.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from algo.core.chain import ChainRow, OptionChainSnapshot
from algo.core.enums import Exchange, Right
from algo.core.errors import DataError
from algo.core.instrument import FutureId, OptionId
from algo.core.quote import Quote
from algo.core.timeutil import ist_to_utc

#: The header row as the scraper writes it, left to right. Verified on load.
EXPECTED_HEADER: tuple[str, ...] = (
    "OI (Lots)",
    "Chng in OI",
    "Volume",
    "LTP",
    "Abs. Chng",
    "Bid Qty",
    "Bid Price",
    "Ask Price",
    "Ask Qty",
    "Strike Price",
    "Bid Qty",
    "Bid Price",
    "Ask Price",
    "Ask Qty",
    "Abs. Chng",
    "LTP",
    "Volume",
    "Chng in OI",
    "OI (Lots)",
)

_STRIKE_COL = 9
#: Column offsets for each side, keyed off the shared header positions above.
_CALL = {"oi": 0, "volume": 2, "ltp": 3, "bid_qty": 5, "bid": 6, "ask": 7, "ask_qty": 8}
_PUT = {"bid_qty": 10, "bid": 11, "ask": 12, "ask_qty": 13, "ltp": 15, "volume": 16, "oi": 18}

_TITLE_RE = re.compile(r"([A-Z0-9]+)\s+Option Chain\s*-\s*(\d{4}-\d{2}-\d{2})")
_ASON_RE = re.compile(r"As on\s+(.+?)\s+IST", re.IGNORECASE)
_UNDERLYING_RE = re.compile(r"Underlying Value:\s*([\d,]+(?:\.\d+)?)")

_ASON_FORMATS = ("%d %b %Y - %H:%M", "%d %B %Y - %H:%M", "%Y-%m-%d %H:%M")


def _decimal(value: Any) -> Decimal | None:
    """A cell as a Decimal, or None when the cell is empty.

    Goes through `str` rather than `Decimal(float)` so a value openpyxl handed
    back as a float does not carry binary-float noise into money arithmetic.
    """
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except InvalidOperation as exc:
        raise DataError(f"not a number: {value!r}") from exc


def _int(value: Any) -> int | None:
    parsed = _decimal(value)
    return int(parsed) if parsed is not None else None


def _verify_header(row: tuple[Any, ...], path: Path) -> None:
    found = tuple("" if cell is None else str(cell).strip() for cell in row[: len(EXPECTED_HEADER)])
    if found != EXPECTED_HEADER:
        raise DataError(
            f"{path}: the option-chain sheet does not have the expected column "
            f"layout. This reader maps columns by position, so it will not guess.\n"
            f"  wanted: {EXPECTED_HEADER}\n"
            f"  found:  {found}"
        )


def _parse_titles(rows: list[tuple[Any, ...]], path: Path) -> tuple[str, date, datetime, Decimal]:
    def cell(i: int) -> str:
        return "" if not rows[i] or rows[i][0] is None else str(rows[i][0]).strip()

    title, ason, underlying = cell(0), cell(1), cell(2)

    match = _TITLE_RE.search(title)
    if match is None:
        raise DataError(f"{path}: could not read symbol and expiry from title line {title!r}")
    symbol, expiry_text = match.group(1), match.group(2)

    match = _ASON_RE.search(ason)
    if match is None:
        raise DataError(f"{path}: could not read the snapshot time from {ason!r}")
    stamp_text = match.group(1).strip()
    stamp = None
    for fmt in _ASON_FORMATS:
        try:
            # Naive on purpose: the page stamps IST wall-clock with no offset,
            # and `ist_to_utc` below is what attaches the zone. Parsing it as
            # aware here would mean inventing a UTC offset the file never gave.
            stamp = datetime.strptime(stamp_text, fmt)  # noqa: DTZ007
            break
        except ValueError:
            continue
    if stamp is None:
        raise DataError(f"{path}: unrecognised snapshot time {stamp_text!r}")

    match = _UNDERLYING_RE.search(underlying)
    if match is None:
        raise DataError(f"{path}: could not read the underlying value from {underlying!r}")
    futures_price = _decimal(match.group(1))
    if futures_price is None or futures_price <= 0:
        raise DataError(f"{path}: underlying value is not a usable price: {underlying!r}")

    # The page stamps IST; everything downstream is UTC.
    ts = ist_to_utc(stamp.date(), stamp.time())
    return symbol, date.fromisoformat(expiry_text), ts, futures_price


def _quote(row: tuple[Any, ...], cols: dict[str, int], ts: datetime) -> Quote:
    return Quote(
        exchange_ts=ts,
        received_ts=ts,
        bid=_decimal(row[cols["bid"]]),
        ask=_decimal(row[cols["ask"]]),
        bid_qty=_int(row[cols["bid_qty"]]),
        ask_qty=_int(row[cols["ask_qty"]]),
        ltp=_decimal(row[cols["ltp"]]),
        volume=_int(row[cols["volume"]]) or 0,
        open_interest=_int(row[cols["oi"]]),
    )


def load_chain(
    path: Path | str,
    *,
    futures_expiry: date | None = None,
    exchange: Exchange = Exchange.MCX,
    sheet: str = "option_chain",
) -> OptionChainSnapshot:
    """Read one scraped live chain into a snapshot.

    `futures_expiry` is the futures contract the options settle into. It is not
    in the scrape — the page shows only the option ladder — so it defaults to the
    option expiry, which is the same documented heuristic
    `bhavcopy.nearest_futures_expiry` applies, and is overridable for the same
    reason (open question Q1c).

    Returns rows with `iv` and `delta` unset. Run the result through
    `algo.pricing.chain_greeks.enrich` to fill them; keeping the two steps apart
    means this module never has to decide a risk-free rate.
    """
    import openpyxl

    path = Path(path)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in workbook.sheetnames:
            raise DataError(f"{path}: no {sheet!r} sheet (found {workbook.sheetnames})")
        rows = list(workbook[sheet].iter_rows(values_only=True))
    finally:
        workbook.close()

    if len(rows) < 6:
        raise DataError(f"{path}: too few rows to be a chain export ({len(rows)})")

    symbol, option_expiry, ts, futures_price = _parse_titles(rows, path)
    _verify_header(rows[4], path)

    future = FutureId(
        underlying=symbol,
        expiry=futures_expiry or option_expiry,
        exchange=exchange,
    )

    chain_rows: list[ChainRow] = []
    for line_no, row in enumerate(rows[5:], start=6):
        if len(row) < len(EXPECTED_HEADER):
            continue
        strike = _decimal(row[_STRIKE_COL])
        if strike is None or strike <= 0:
            continue
        for right, cols in ((Right.CE, _CALL), (Right.PE, _PUT)):
            try:
                quote = _quote(row, cols, ts)
            except DataError as exc:
                raise DataError(f"{path}:{line_no} {right.value} {strike}: {exc}") from exc
            chain_rows.append(
                ChainRow(
                    option=OptionId(
                        underlying_future=future,
                        option_expiry=option_expiry,
                        strike=strike,
                        right=right,
                        exchange=exchange,
                    ),
                    quote=quote,
                )
            )

    if not chain_rows:
        raise DataError(f"{path}: header parsed but no strike rows followed")

    chain_rows.sort(key=lambda r: (r.strike, r.right.value))
    return OptionChainSnapshot(
        ts=ts,
        underlying=symbol,
        option_expiry=option_expiry,
        futures_price=futures_price,
        rows=tuple(chain_rows),
    )
