"""Poll live MCX futures quotes via Kotak Neo and log them to an Excel workbook.

Standalone from the trading engine: read-only market data, no orders, no
strategy. Each poll appends one row to an .xlsx log so a history builds up
across runs. Meant to be driven by a scheduler — with no flags it does one
poll and exits, which is what Windows Task Scheduler should call every few
minutes; pass ``--loop`` instead for an interactive session that polls
itself on an interval.

Needs only ``ALGO_KOTAK_CONSUMER_KEY`` (or ``ALGO_KOTAK_MARKET_DATA_KEY``) in
the environment. Quotes authenticate on the consumer key alone — no TOTP/MPIN
session — see algo/data/kotak_feed.py.

This repo has no sourced MCX holiday calendar yet (algo/exchange/calendar.py
only ships a synthetic one), so `market_is_open` checks weekday + session
hours only. A poll that lands on an unlisted holiday just gets back a stale
last-traded price rather than a fresh one.
"""

from __future__ import annotations

import argparse
import sys
import time
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from structlog import get_logger

from algo.core.enums import Exchange
from algo.core.timeutil import is_us_dst
from algo.data.kotak_feed import NeoQuotesTransport
from algo.exchange.master import (
    InstrumentMaster,
    KotakMasterSource,
    fetch_master,
)
from algo.execution.kotak import credentials_from_env

log = get_logger(__name__)
IST = ZoneInfo("Asia/Kolkata")
SHEET_NAME = "live_quotes"
HEADERS = [
    "polled_at_ist",
    "symbol",
    "tradingsymbol",
    "expiry",
    "ltp",
    "bid",
    "ask",
    "volume",
    "open_interest",
]


def market_is_open(now_ist: datetime) -> bool:
    if now_ist.weekday() >= 5:
        return False
    close_hour, close_minute = (23, 30) if is_us_dst(now_ist.date()) else (23, 55)
    open_t = now_ist.replace(hour=9, minute=0, second=0, microsecond=0)
    close_t = now_ist.replace(hour=close_hour, minute=close_minute, second=0, microsecond=0)
    return open_t <= now_ist < close_t


def ensure_workbook(path: Path) -> None:
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = SHEET_NAME
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(name="Arial", bold=True)
    workbook.save(path)


def append_row(path: Path, row: list[object]) -> None:
    workbook = load_workbook(path)
    sheet = workbook[SHEET_NAME]
    sheet.append(row)
    for cell in sheet[sheet.max_row]:
        cell.font = Font(name="Arial")
    workbook.save(path)


def _payload_for(payloads: list[dict[str, Any]], token: str) -> dict[str, Any] | None:
    """Match a quote payload to its token, tolerating the `SEGMENT|token` prefix
    the API sometimes writes into `exchange_token`."""
    for entry in payloads:
        exchange_token = str(entry.get("exchange_token") or "")
        if exchange_token == token or exchange_token.endswith(f"|{token}"):
            return entry
    return None


def _num(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def poll_once(
    *,
    symbol: str,
    consumer_key: str,
    master_path: Path,
    out_path: Path,
    refresh_master: bool,
) -> bool:
    """One poll. Returns True if a row was written, False if skipped."""
    now_ist = datetime.now(tz=IST)
    if not market_is_open(now_ist):
        log.info("market closed, skipping poll", at=now_ist.strftime("%Y-%m-%d %H:%M:%S"))
        return False

    if refresh_master or not master_path.exists():
        log.info("fetching Kotak instrument master", path=str(master_path))
        master = fetch_master(
            KotakMasterSource(consumer_key=consumer_key), master_path, now=now_ist
        )
    else:
        master = InstrumentMaster.from_snapshot(master_path)

    futures = master.future_rows(symbol, Exchange.MCX)
    if not futures:
        log.warning("no futures contract in master snapshot", symbol=symbol, path=str(master_path))
        return False
    row = futures[0]  # nearest expiry

    transport = NeoQuotesTransport(consumer_key)
    payload = transport.quotes(
        [{"exchange_segment": "mcx_fo", "instrument_token": row.symboltoken}]
    )
    if isinstance(payload, dict):
        log.warning(
            "quote poll failed", error=payload.get("Error") or payload.get("error") or payload
        )
        return False
    if not isinstance(payload, list):
        log.warning("quote poll returned an unreadable payload", payload=repr(payload))
        return False
    payloads = [entry for entry in payload if isinstance(entry, dict)]
    quote = _payload_for(payloads, row.symboltoken)
    if quote is None:
        log.warning("no quote returned", tradingsymbol=row.tradingsymbol, token=row.symboltoken)
        return False

    ensure_workbook(out_path)
    append_row(
        out_path,
        [
            now_ist.strftime("%Y-%m-%d %H:%M:%S"),
            symbol,
            row.tradingsymbol,
            row.expiry.isoformat() if row.expiry else "",
            float(_num(quote.get("ltp") or quote.get("last_traded_price")) or 0),
            float(_num(quote.get("bid")) or 0) or None,
            float(_num(quote.get("ask")) or 0) or None,
            int(_num(quote.get("volume") or quote.get("vol")) or 0),
            int(_num(quote.get("open_int") or quote.get("openInterest")) or 0),
        ],
    )
    log.info("quote logged", tradingsymbol=row.tradingsymbol, out=str(out_path))
    return True


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--symbol", default="GOLDM", help="MCX underlying (default GOLDM)")
    parser.add_argument(
        "--out", type=Path, default=Path("data/mcx_live.xlsx"), help="Excel log to append to"
    )
    parser.add_argument(
        "--master",
        type=Path,
        default=Path("state/kotak_master.json"),
        help="Kotak instrument-master snapshot (fetched automatically if missing)",
    )
    parser.add_argument(
        "--refresh-master", action="store_true", help="Re-download the instrument master first"
    )
    parser.add_argument(
        "--loop", action="store_true", help="Keep running, polling every --interval-minutes"
    )
    parser.add_argument(
        "--interval-minutes", type=float, default=5.0, help="Poll interval when --loop is set"
    )
    args = parser.parse_args()

    load_dotenv()
    credentials = credentials_from_env()
    consumer_key = credentials.market_data_key or credentials.consumer_key
    if not consumer_key:
        log.error(
            "no Kotak consumer key set",
            hint="set ALGO_KOTAK_CONSUMER_KEY (or ALGO_KOTAK_MARKET_DATA_KEY); "
            "copy .env.example to .env and fill it in",
        )
        raise SystemExit(1)

    if not args.loop:
        poll_once(
            symbol=args.symbol,
            consumer_key=consumer_key,
            master_path=args.master,
            out_path=args.out,
            refresh_master=args.refresh_master,
        )
        return

    log.info("looping", interval_minutes=args.interval_minutes, stop="Ctrl+C")
    refresh = args.refresh_master
    try:
        while True:
            poll_once(
                symbol=args.symbol,
                consumer_key=consumer_key,
                master_path=args.master,
                out_path=args.out,
                refresh_master=refresh,
            )
            refresh = False  # only refresh once per loop unless the file goes missing
            time.sleep(args.interval_minutes * 60)
    except KeyboardInterrupt:
        log.info("stopped")


if __name__ == "__main__":
    main()
