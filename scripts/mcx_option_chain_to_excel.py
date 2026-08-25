"""Snapshot the live MCX GOLDM option chain via Kotak Neo into an Excel sheet
laid out like MCX's own option-chain view: calls on the left, strike in the
middle, puts on the right, one row per strike.

Standalone from the trading engine: read-only market data, no orders, no
strategy. Unlike mcx_live_to_excel.py (which appends a time series of one
instrument), this **overwrites** the sheet every run — a chain is a snapshot,
not a series. Meant to be driven by a scheduler; with no flags it does one
poll and exits.

Needs only ``ALGO_KOTAK_CONSUMER_KEY`` (or ``ALGO_KOTAK_MARKET_DATA_KEY``) in
the environment — quotes authenticate on the consumer key alone, no
TOTP/MPIN session (see algo/data/kotak_feed.py).

**Chng in OI is a session proxy, not the exchange figure.** Kotak's quotes
API returns the *current* open interest but no previous-close OI or an
OI-delta field, so the true "change since yesterday's close" the MCX website
shows cannot be reproduced from a live quote alone (it would need yesterday's
bhavcopy — see algo/data/bhavcopy.py — which is not fetched here). Instead
this script remembers each contract's open interest the first time it is
polled each calendar day (state/mcx_oi_baseline_<date>.json) and reports the
change from *that*, i.e. change-since-this-script-started-watching-today.
Restarting the scheduled task mid-session resets the baseline.

Certificates: see the note in mcx_live_to_excel.py — `truststore` is
required for the same reason.
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import truststore

truststore.inject_into_ssl()

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from structlog import get_logger  # noqa: E402

from algo.core.enums import Exchange  # noqa: E402
from algo.core.timeutil import is_us_dst  # noqa: E402
from algo.data.kotak_feed import MAX_TOKENS_PER_CALL, NeoQuotesTransport  # noqa: E402
from algo.exchange.master import (  # noqa: E402
    InstrumentMaster,
    KotakMasterSource,
    MasterRow,
    fetch_master,
)
from algo.execution.kotak import credentials_from_env  # noqa: E402

log = get_logger(__name__)
IST = ZoneInfo("Asia/Kolkata")
SHEET_NAME = "option_chain"

#: Column order, left (calls) to right (puts), matching the MCX option-chain view.
CALL_COLUMNS = ["OI (Lots)", "Chng in OI", "Volume", "LTP", "Abs. Chng", "Bid Qty", "Bid Price"]
CALL_TAIL = ["Ask Price", "Ask Qty"]
PUT_HEAD = ["Bid Qty", "Bid Price", "Ask Price", "Ask Qty"]
PUT_COLUMNS = ["Abs. Chng", "LTP", "Volume", "Chng in OI", "OI (Lots)"]


def market_is_open(now_ist: datetime) -> bool:
    if now_ist.weekday() >= 5:
        return False
    close_hour, close_minute = (23, 30) if is_us_dst(now_ist.date()) else (23, 55)
    open_t = now_ist.replace(hour=9, minute=0, second=0, microsecond=0)
    close_t = now_ist.replace(hour=close_hour, minute=close_minute, second=0, microsecond=0)
    return open_t <= now_ist < close_t


def _num(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _best_price(quote: dict[str, Any], side: str) -> Decimal | None:
    nested = quote.get("depth")
    levels = nested.get(side) if isinstance(nested, dict) else None
    if isinstance(levels, list) and levels and isinstance(levels[0], dict):
        price = _num(levels[0].get("price"))
        if price is not None and price > 0:
            return price
    return None


def _best_qty(quote: dict[str, Any], side: str) -> int | None:
    nested = quote.get("depth")
    levels = nested.get(side) if isinstance(nested, dict) else None
    if isinstance(levels, list) and levels and isinstance(levels[0], dict):
        qty = _num(levels[0].get("quantity"))
        if qty is not None:
            return int(qty)
    return None


def _payload_for(payloads: list[dict[str, Any]], token: str) -> dict[str, Any] | None:
    for entry in payloads:
        exchange_token = str(entry.get("exchange_token") or "")
        if exchange_token == token or exchange_token.endswith(f"|{token}"):
            return entry
    return None


def fetch_quotes(
    transport: NeoQuotesTransport, rows: list[MasterRow]
) -> dict[str, dict[str, Any]]:
    """One `payloads` dict keyed by symboltoken, chunked under the API's limit."""
    by_token: dict[str, dict[str, Any]] = {}
    for start in range(0, len(rows), MAX_TOKENS_PER_CALL):
        chunk = rows[start : start + MAX_TOKENS_PER_CALL]
        tokens = [
            {"exchange_segment": "mcx_fo", "instrument_token": r.symboltoken} for r in chunk
        ]
        payload = transport.quotes(tokens)
        if not isinstance(payload, list):
            log.warning("quote chunk failed", start=start, payload=repr(payload)[:300])
            continue
        for entry in payload:
            if not isinstance(entry, dict):
                continue
            for row in chunk:
                if _payload_for([entry], row.symboltoken) is not None:
                    by_token[row.symboltoken] = entry
    return by_token


class OiBaseline:
    """Each contract's first-seen-today open interest, so `Chng in OI` has
    something to compare against (see the module docstring's caveat)."""

    def __init__(self, path: Path, today: date) -> None:
        self._path = path
        self._today = today.isoformat()
        self._data: dict[str, Any] = {}
        if path.exists():
            raw = json.loads(path.read_text(encoding="utf-8"))
            if raw.get("date") == self._today:
                self._data = raw.get("oi", {})

    def delta(self, tradingsymbol: str, current_oi: int | None) -> int | None:
        if current_oi is None:
            return None
        if tradingsymbol not in self._data:
            self._data[tradingsymbol] = current_oi
            return 0
        return current_oi - int(self._data[tradingsymbol])

    def save(self) -> None:
        self._path.parent.mkdir(parents=True, exist_ok=True)
        self._path.write_text(
            json.dumps({"date": self._today, "oi": self._data}, indent=2), encoding="utf-8"
        )


def _side_metrics(
    row: MasterRow, quote: dict[str, Any] | None, lot_size: int, baseline: OiBaseline
) -> dict[str, Any]:
    if quote is None:
        return {
            "oi": None, "chng_oi": None, "volume": None, "ltp": None, "abs_chng": None,
            "bid_qty": None, "bid": None, "ask": None, "ask_qty": None,
        }
    oi_units = _num(quote.get("open_int"))
    oi_lots = int(oi_units / lot_size) if oi_units is not None else None
    volume_units = _num(quote.get("last_volume"))
    volume_lots = int(volume_units / lot_size) if volume_units is not None else None
    ltp = _num(quote.get("ltp"))
    return {
        "oi": oi_lots,
        "chng_oi": baseline.delta(row.tradingsymbol, oi_lots),
        "volume": volume_lots,
        "ltp": float(ltp) if ltp is not None else None,
        "abs_chng": float(_num(quote.get("change")) or 0) if ltp is not None else None,
        "bid_qty": _best_qty(quote, "buy"),
        "bid": _to_float(_best_price(quote, "buy")),
        "ask": _to_float(_best_price(quote, "sell")),
        "ask_qty": _best_qty(quote, "sell"),
    }


def _to_float(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


def build_workbook(
    *,
    symbol: str,
    expiry: date,
    strikes: list[Decimal],
    calls: dict[Decimal, dict[str, Any]],
    puts: dict[Decimal, dict[str, Any]],
    underlying_ltp: float | None,
    as_of: datetime,
) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = SHEET_NAME

    bold = Font(name="Arial", bold=True)
    plain = Font(name="Arial")
    center = Alignment(horizontal="center")

    sheet["A1"] = f"{symbol} Option Chain - {expiry.isoformat()}"
    sheet["A1"].font = Font(name="Arial", bold=True, size=14)
    sheet["A2"] = f"As on {as_of:%d %b %Y - %H:%M} IST"
    sheet["A2"].font = plain
    if underlying_ltp is not None:
        sheet["A3"] = f"Underlying Value: {underlying_ltp:,.2f}"
        sheet["A3"].font = bold

    header_row = 5
    strike_col = len(CALL_COLUMNS) + len(CALL_TAIL) + 1
    sheet.cell(row=header_row - 1, column=1, value="CALLS").font = bold
    sheet.cell(row=header_row - 1, column=strike_col + 1, value="PUTS").font = bold

    headers = CALL_COLUMNS + CALL_TAIL + ["Strike Price"] + PUT_HEAD + PUT_COLUMNS
    for col, name in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col, value=name)
        cell.font = bold
        cell.alignment = center

    for r, strike in enumerate(strikes, start=header_row + 1):
        call = calls.get(strike, {})
        put = puts.get(strike, {})
        values = [
            call.get("oi"), call.get("chng_oi"), call.get("volume"), call.get("ltp"),
            call.get("abs_chng"), call.get("bid_qty"), call.get("bid"),
            call.get("ask"), call.get("ask_qty"),
            float(strike),
            put.get("bid_qty"), put.get("bid"), put.get("ask"), put.get("ask_qty"),
            put.get("abs_chng"), put.get("ltp"), put.get("volume"),
            put.get("chng_oi"), put.get("oi"),
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=r, column=col, value=value)
            cell.font = plain
            if col == strike_col + 1:
                cell.font = bold

    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 12

    return workbook


def poll_once(
    *,
    symbol: str,
    consumer_key: str,
    master_path: Path,
    out_path: Path,
    baseline_dir: Path,
    expiry_override: date | None,
    refresh_master: bool,
) -> bool:
    now_ist = datetime.now(tz=IST)
    if not market_is_open(now_ist):
        log.info("market closed, skipping poll", at=now_ist.strftime("%Y-%m-%d %H:%M:%S"))
        return False

    if refresh_master or not master_path.exists():
        log.info("fetching Kotak instrument master", path=str(master_path))
        master = fetch_master(
            KotakMasterSource(consumer_key=consumer_key), master_path, now=now_ist
        )
    else:
        master = InstrumentMaster.from_snapshot(master_path)

    expiries = master.option_expiries(symbol, Exchange.MCX)
    if not expiries:
        log.warning("no option expiries listed", symbol=symbol)
        return False
    expiry = expiry_override or expiries[0]

    option_rows = master.option_rows(symbol, Exchange.MCX, expiry)
    if not option_rows:
        log.warning("no option rows for expiry", symbol=symbol, expiry=str(expiry))
        return False
    lot_size = int(option_rows[0].lot_size or 100)

    futures = master.future_rows(symbol, Exchange.MCX)
    transport = NeoQuotesTransport(consumer_key)

    underlying_ltp: float | None = None
    if futures:
        fut_payload = transport.quotes(
            [{"exchange_segment": "mcx_fo", "instrument_token": futures[0].symboltoken}]
        )
        if isinstance(fut_payload, list) and fut_payload:
            quote = _payload_for(
                [e for e in fut_payload if isinstance(e, dict)], futures[0].symboltoken
            )
            if quote is not None:
                underlying_ltp = _to_float(_num(quote.get("ltp")))

    by_token = fetch_quotes(transport, list(option_rows))
    if not by_token:
        log.warning("no option quotes returned at all", symbol=symbol, expiry=str(expiry))
        return False

    baseline_path = baseline_dir / f"mcx_oi_baseline_{now_ist.date().isoformat()}.json"
    baseline = OiBaseline(baseline_path, now_ist.date())

    calls: dict[Decimal, dict[str, Any]] = {}
    puts: dict[Decimal, dict[str, Any]] = {}
    strikes: set[Decimal] = set()
    for row in option_rows:
        if row.strike is None:
            continue
        right = "CE" if row.tradingsymbol.upper().endswith("CE") else "PE"
        quote = by_token.get(row.symboltoken)
        metrics = _side_metrics(row, quote, lot_size, baseline)
        strikes.add(row.strike)
        (calls if right == "CE" else puts)[row.strike] = metrics
    baseline.save()

    ensure_parent(out_path)
    workbook = build_workbook(
        symbol=symbol,
        expiry=expiry,
        strikes=sorted(strikes),
        calls=calls,
        puts=puts,
        underlying_ltp=underlying_ltp,
        as_of=now_ist,
    )
    workbook.save(out_path)
    log.info(
        "option chain logged",
        symbol=symbol,
        expiry=str(expiry),
        strikes=len(strikes),
        out=str(out_path),
    )
    return True


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--symbol", default="GOLDM", help="MCX underlying (default GOLDM)")
    parser.add_argument(
        "--expiry",
        type=str,
        default=None,
        help="Option expiry YYYY-MM-DD (default: nearest listed)",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=Path("data/mcx_option_chain.xlsx"),
        help="Excel file to (re)write with the current chain snapshot",
    )
    parser.add_argument(
        "--master",
        type=Path,
        default=Path("state/kotak_master.json"),
        help="Kotak instrument-master snapshot (fetched automatically if missing)",
    )
    parser.add_argument(
        "--baseline-dir",
        type=Path,
        default=Path("state"),
        help="Where the daily OI baseline cache is kept",
    )
    parser.add_argument(
        "--refresh-master", action="store_true", help="Re-download the instrument master first"
    )
    parser.add_argument(
        "--loop", action="store_true", help="Keep running, polling every --interval-minutes"
    )
    parser.add_argument(
        "--interval-minutes", type=float, default=5.0, help="Poll interval when --loop is set"
    )
    args = parser.parse_args()

    load_dotenv()
    credentials = credentials_from_env()
    consumer_key = credentials.market_data_key or credentials.consumer_key
    if not consumer_key:
        log.error(
            "no Kotak consumer key set",
            hint="set ALGO_KOTAK_CONSUMER_KEY (or ALGO_KOTAK_MARKET_DATA_KEY); "
            "copy .env.example to .env and fill it in",
        )
        raise SystemExit(1)

    expiry_override = date.fromisoformat(args.expiry) if args.expiry else None

    def run_once() -> None:
        poll_once(
            symbol=args.symbol,
            consumer_key=consumer_key,
            master_path=args.master,
            out_path=args.out,
            baseline_dir=args.baseline_dir,
            expiry_override=expiry_override,
            refresh_master=args.refresh_master,
        )

    if not args.loop:
        run_once()
        return

    log.info("looping", interval_minutes=args.interval_minutes, stop="Ctrl+C")
    try:
        while True:
            run_once()
            time.sleep(args.interval_minutes * 60)
    except KeyboardInterrupt:
        log.info("stopped")


if __name__ == "__main__":
    main()
